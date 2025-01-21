import openpyxl
import os


class RegisterUserTestData:

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        current_dir = os.path.dirname(os.path.abspath(__file__))
        excel_path = os.path.join(current_dir, '../TestData/Testdata.xlsx')
        book = openpyxl.load_workbook(excel_path)
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]

